import os, time, sys, json, shutil, traceback, logging
import pyautogui as pg
import tkinter as tk
import openpyxl as xl
import pandas as pd

from datetime import datetime
from logging.handlers import RotatingFileHandler
from logging import Logger
from openpyxl.cell.cell import Cell
from openpyxl.styles import PatternFill
from typing import Tuple
from pathlib import Path
from tkinter import filedialog, messagebox, simpledialog
from openpyxl.worksheet.worksheet import Worksheet


WINDOWS_VER  = 11 if sys.getwindowsversion().build >= 22000 else 10
WINDOWS_PATH = "win"+str(WINDOWS_VER)


def get_exe_dir() -> str:
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    else:
        return os.path.dirname(os.path.abspath(__file__))


def get_logger() -> Logger:
    log_path = os.path.join(get_exe_dir(), "gutterdone.log")
    handler = RotatingFileHandler(log_path, maxBytes=1_000_000, backupCount=3)
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    handler.setFormatter(formatter)

    logger = logging.getLogger()
    logger.setLevel(logging.INFO)
    logger.addHandler(handler)
    logger.addHandler(logging.StreamHandler())

    return logger


def get_resource_path(relative_path) -> str:
    """Get the path to a resource, whether running as script or exe."""
    try:
        base_path = sys._MEIPASS  # Set by PyInstaller
    except AttributeError:
        base_path = os.path.abspath(".")
        
    return os.path.join(base_path, relative_path)


def wait_for(
        image: str,
        confidence: float = 0.8,
        timeout: int = 10,
    ) -> tuple | bool:
    """
    Waits until specified image appears on screen before returning its coordinates

    Args:
        image (str): Path to the image you are searching for on screen.
        confidence (float): How confident the opencv model most be in order to return coordinates for your iamge.
        timeout (int): How long to search for before timeout.

    Returns:
        tuple: The bounding box for your image on screen.
    """

    coords = None
    start = time.time()

    while coords is None:
        try:
            coords = pg.locateOnScreen(image, confidence=confidence)
            time.sleep(0.3)

            if coords:
                time.sleep(0.3)
                return coords
            
        except pg.ImageNotFoundException:
            pass

        if time.time() - start > timeout:
            raise TimeoutError(f"Timed out waiting for image: {image}")


def get_coords(box, x_ratio: float, y_ratio: float) -> tuple:
    """
    Returns coordinates to click on based on image and ratios. 
    Specifying the ratio moves the point down in the Y direction and rightwards in the X direction

    Args:
        box (tuple): The area of the screen occupied by your image.
        x_ratio (float): The percentage/fraction of the image to the right you intend to click.
        y_ratio (float): The percentage/fraction of the image down you intend to click.
    
    Returns:
        tuple: The x,y coordinates to click.
    """
    x, y, w, h = box

    return (int(x + (x_ratio * w)), int(y + (y_ratio * h)))


def prepare_express(folder):
    """
    Sets file save destination at beginning of program run.
    
    Args:
        folder: Folder path your selected excel lives in.
    """

    pg.click(960, 540) # click center screen to change focus

    pg.click(pg.center(wait_for(get_resource_path(f"images/{WINDOWS_PATH}/FILE.PNG"))))
    pg.click(pg.center(wait_for(get_resource_path(f"images/{WINDOWS_PATH}/SAVE_AS.PNG"))))

    pg.write("init_file", interval=0.01)
    time.sleep(0.5)

    pg.click(pg.center(wait_for(get_resource_path(f"images/{WINDOWS_PATH}/FOLDER_PATH.PNG"))))

    pg.write(folder, interval=0.01)
    pg.press("enter")

    pg.click(pg.center(wait_for(get_resource_path(f"images/{WINDOWS_PATH}/SAVE_FILE.PNG"))))
    pg.press("enter")


def input_values(throat, slope, total_q, inlet):
    exn_coords = get_coords(wait_for(get_resource_path(f"images/{WINDOWS_PATH}/EXPRESS_NAME.PNG")), 1/4, 1/2)
    pg.click(*exn_coords) # EXPRESS_NAME.PNG
    pg.hotkey("ctrl", "a")
    pg.press("backspace")
    pg.write(inlet, interval=0.01)

    sl_coords = get_coords(wait_for(get_resource_path(f"images/{WINDOWS_PATH}/THROAT.PNG")), 4/3, 1/2)
    pg.click(*sl_coords)
    pg.write(str(throat), interval=0.01)

    sl_coords = get_coords(wait_for(get_resource_path(f"images/{WINDOWS_PATH}/SLOPE.PNG")), 7/8, 1/2)
    pg.click(*sl_coords)
    pg.write(str(slope), interval=0.01)

    q_coords = get_coords(wait_for(get_resource_path(f"images/{WINDOWS_PATH}/Q.PNG")), 7/8, 3/4)
    pg.click(*q_coords)
    pg.write(str(total_q), interval=0.01)


def run_express(inlet: str):
    pg.click(pg.center(wait_for(get_resource_path(f"images/{WINDOWS_PATH}/RUN.PNG"))))

    save_hxp(inlet)
    time.sleep(0.5)

    save_pdf(inlet)
    time.sleep(0.5)

    save_csv(inlet)
    time.sleep(0.5)


def save_hxp(inlet: str):
    pg.click(pg.center(wait_for(get_resource_path(f"images/{WINDOWS_PATH}/FILE.PNG"))))
    pg.click(pg.center(wait_for(get_resource_path(f"images/{WINDOWS_PATH}/SAVE_AS.PNG"))))
    time.sleep(2)

    pg.write(inlet, interval=0.01)
    time.sleep(0.5)

    pg.press("enter")
    pg.press("enter")


def save_pdf(inlet: str):
    pg.click(pg.center(wait_for(get_resource_path(f"images/{WINDOWS_PATH}/PRINT.PNG"))))
    pg.click(pg.center(wait_for(get_resource_path(f"images/{WINDOWS_PATH}/REPORT.PNG"))))

    if WINDOWS_VER == 11:
        wait_for(get_resource_path(f"images/{WINDOWS_PATH}/PRINT_SCREEN.PNG"))
        
        pg.press("enter")
        pg.write("Microsoft Print to PDF", interval=0.01)
        pg.press("enter")

        pb_coords = get_coords(wait_for(get_resource_path(f"images/{WINDOWS_PATH}/PRINT_IN_PRINT_SCREEN.PNG")), 1/4, 1/2)
        pg.click(*pb_coords)

        wait_for(get_resource_path(f"images/{WINDOWS_PATH}/PRINT_FILE_EXPLORER.PNG"))
        pg.write(inlet, interval=0.01)

        time.sleep(0.5)
        pg.press("enter")

        pc_coords = get_coords(wait_for(get_resource_path(f"images/{WINDOWS_PATH}/PRINTING_COMPLETE.PNG")), 3/4, 5/6)
        pg.click(*pc_coords)

    if WINDOWS_VER == 10:
        wait_for(get_resource_path(f"images/{WINDOWS_PATH}/PRINT_SCREEN.PNG"))
        pg.write("Microsoft Print to PDF", interval=0.01)
        pg.press("enter")

        time.sleep(0.5)
        pg.write(inlet, interval=0.01)

        time.sleep(0.5)
        pg.press("enter", 2, 1)


def save_csv(inlet: str):
    pg.click(960, 540)
    pg.click(pg.center(wait_for(get_resource_path(f"images/{WINDOWS_PATH}/FILE.PNG"))))

    pg.click(pg.center(wait_for(get_resource_path(f"images/{WINDOWS_PATH}/EXPORT.PNG"))))
    pg.click(pg.center(wait_for(get_resource_path(f"images/{WINDOWS_PATH}/RESULTS_GRID.PNG"))))
    pg.click(pg.center(wait_for(get_resource_path(f"images/{WINDOWS_PATH}/CSV.PNG"))))

    wait_for(get_resource_path(f"images/{WINDOWS_PATH}/CSV_FILE_EXPLORER.PNG"))
    pg.write(inlet.strip() + ".csv", interval=0.01)
    time.sleep(0.5)

    pg.press("enter")


def read_csv(folder: str, inlet: str) -> pd.DataFrame:
    file = (inlet.strip()) + ".csv"
    path = Path(folder) / file

    df = pd.read_csv(path, header=1)
    df = df.loc[[1]].reset_index(drop=True)

    return df


def input_sag(row: Tuple[Cell, ...], prev_carryover_q: float):
    q = float(row[XL_INDEX["Q"]].value)

    total_q = q + prev_carryover_q
    if total_q >= 10.62:
        highlight_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        row[XL_INDEX["INTERCEPTED_FLOW"]].fill = highlight_fill

    row[XL_INDEX["CARRYOVER_Q"]].value = float(prev_carryover_q)
    row[XL_INDEX["TOTAL_Q"]].value     = total_q

    row[XL_INDEX["INTERCEPTED_FLOW"]].value = total_q
    row[XL_INDEX["Q_BYPASS"]].value         = "N/A"
    row[XL_INDEX["SPREAD"]].value           = "N/A"
    row[XL_INDEX["DEPTH"]].value            = "N/A"


def edit_xl(row: Tuple[Cell, ...], df_express: pd.DataFrame, prev_carryover_q: float):
    row[XL_INDEX["CARRYOVER_Q"]].value = prev_carryover_q
    row[XL_INDEX["TOTAL_Q"]].value     = prev_carryover_q + float(row[XL_INDEX["Q"]].value)

    row[XL_INDEX["INTERCEPTED_FLOW"]].value = float(df_express.at[0, "Captured"])
    row[XL_INDEX["Q_BYPASS"]].value         = float(df_express.at[0, "Q"])
    row[XL_INDEX["SPREAD"]].value           = float(df_express.at[0, "Spread"])
    row[XL_INDEX["DEPTH"]].value            = float(df_express.at[0, "Depth"])

    if float(df_express.at[0, "Spread"]) > 6.2:
        highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        row[XL_INDEX["SPREAD"]].fill = highlight_fill


def iter_xl(ws: Worksheet, folder, start_row: int):
    prev_carryover_q = 0
    prev_to_inlet = ""

    for row in ws.iter_rows(min_row=start_row):
        inlet_cell    = row[XL_INDEX["INLET"]]
        q_cell        = row[XL_INDEX["Q"]]
        long_cell     = row[XL_INDEX["LONG"]]
        type_cell     = row[XL_INDEX["ON-GRADE/SAG"]]
        throat_cell   = row[XL_INDEX["THROAT"]]
        to_inlet_cell = row[XL_INDEX["TO_INLET"]]

        inlet: str    = inlet_cell.value
        type: str     = type_cell.value
        throat: str   = throat_cell.value
        to_inlet: str = to_inlet_cell.value
        q: float      = q_cell.value
        long: float   = long_cell.value

        if type == "Sag":
            input_sag(row, prev_carryover_q)
            continue

        if prev_to_inlet == inlet:
            total_q = q + prev_carryover_q
        else:
            total_q = q

        long_percent = long * 100
        input_values(throat, long_percent, total_q, inlet)
        run_express(inlet)

        df_express = read_csv(folder, inlet)
        edit_xl(row, df_express, prev_carryover_q)

        prev_carryover_q = float(df_express.at[0, "Q"])
        prev_to_inlet    = to_inlet


def mkdirs(folder):
    dirs = ["/csvs", "/pdfs", "/hxps"]
    for dir in dirs:
        try:
            os.makedirs(folder+dir)
        except FileExistsError:
            shutil.rmtree(folder+dir)
            os.makedirs(folder+dir)


def move_files(folder):

    for file in os.listdir(folder):

        filename = os.fsdecode(file)
        original_path = Path(folder) / filename

        if filename.endswith(".pdf"):
            new_path = Path(folder) / "pdfs" / filename
            os.rename(original_path, new_path)

        elif filename.endswith(".hxp"):
            new_path = Path(folder) / "hxps" / filename
            os.rename(original_path, new_path)

        elif filename.endswith(".csv"):
            new_path = Path(folder) / "csvs" / filename
            os.rename(original_path, new_path)


if __name__ == "__main__":
    root = tk.Tk()
    root.withdraw()

    now = datetime.now()
    formatted_datetime = now.strftime("%Y-%m-%d %H-%M-%S")

    logger = get_logger()

    try:
        logger.info("GutterDone Started.")

        screen_w, screen_h = pg.size()
        if screen_w != 1920 or screen_h != 1080:
            raise RuntimeError("Screen resolution must be 1920x1080.")

        result = messagebox.askyesno("Confirmation", "Use default config file?")
        if result:
            with open(get_resource_path('config.json'), 'r') as file:
                XL_INDEX = json.load(file)

        else:
            config_file_path = filedialog.askopenfilename(title="Select Excel Header Config File")
            with open(config_file_path, 'r') as file:
                XL_INDEX = json.load(file)


        xl_file_path = filedialog.askopenfilename(title="Select Gutter Spread Excel")
        folder = os.path.dirname(xl_file_path)

        filename, file_extension = os.path.splitext(xl_file_path)
        if file_extension != ".xlsx":
            raise ValueError("Incorrect file type. File must have .xlsx file extension")
        
        start_row = simpledialog.askinteger("Start Row", "Enter the starting row number:")
        if start_row is None or start_row <= 0:
            raise ValueError("Not a valid start row number: ", start_row)

        prepare_express(folder)

        wb = xl.load_workbook(xl_file_path)
        ws = wb.worksheets[0]

        iter_xl(ws, folder, start_row)
        mkdirs(folder)
        move_files(folder)

        path = Path(folder) / f"GUTTER_SPREAD_EXCEL - {formatted_datetime}.xlsx"
        wb.save(path)

        logger.info("GutterDone finished successfully.")

    except FileNotFoundError as e:
        logger.error("GutterDone failed:")
        logger.exception(e)

    except Exception as e:
        logger.error("GutterDone failed:")
        logger.exception(e)
        
        path = Path(folder) / f"FAILED_GUTTER_SPREAD_EXCEL - {formatted_datetime}.xlsx"
        wb.save(path)

