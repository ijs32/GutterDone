# IN ORDER TO USE THIS TOOL YOU MUST FULL SCREEN HYDRAFLOW EXPRESS AND PUT IT ON YOUR PRIMARY DISPLAY WITH DIMENSIONS 1920x1080
import os, time
import pyautogui as pg
import tkinter as tk
import openpyxl as xl
import pandas as pd

from openpyxl.cell.cell import Cell
from openpyxl.styles import PatternFill
from typing import Tuple
from pathlib import Path
from tkinter import filedialog
from openpyxl.worksheet.worksheet import Worksheet

from config import XL_INDEX, EXPRESS_COORDS


def wait_for(image: str, timeout: int = 10):
    pass


def prepare_express(folder):
    pg.click(x=EXPRESS_COORDS["CENTER"]["x"], y=EXPRESS_COORDS["CENTER"]["y"])

    pg.click(x=EXPRESS_COORDS["FILE"]["x"], y=EXPRESS_COORDS["FILE"]["y"]) # FILE.PNG
    time.sleep(1)

    pg.click(x=EXPRESS_COORDS["SAVE_AS"]["x"], y=EXPRESS_COORDS["SAVE_AS"]["y"]) # SAVE_AS.PNG
    time.sleep(1)

    pg.write("init_file", interval=0.01)
    time.sleep(1)

    pg.click(x=EXPRESS_COORDS["FOLDER_PATH"]["x"], y=EXPRESS_COORDS["FOLDER_PATH"]["y"]) # FOLDER_PATH.PNG
    time.sleep(1)

    pg.write(folder, interval=0.01)
    pg.press("enter")
    time.sleep(1)

    pg.click(x=EXPRESS_COORDS["SAVE_FILE"]["x"], y=EXPRESS_COORDS["SAVE_FILE"]["y"]) # SAVE_FILE.PNG
    time.sleep(1)

    pg.press("enter")
    time.sleep(1)

    pg.click(x=EXPRESS_COORDS["INLETS"]["x"], y=EXPRESS_COORDS["INLETS"]["y"]) # INLETS.PNG
    time.sleep(1)

    pg.click(x=EXPRESS_COORDS["LOCATION"]["x"], y=EXPRESS_COORDS["LOCATION"]["y"]) # LOCATION.PNG
    time.sleep(1)

    pg.click(x=EXPRESS_COORDS["ON_GRADE"]["x"], y=EXPRESS_COORDS["ON_GRADE"]["y"]) # ON_GRADE_DROPDOWN.PNG
    time.sleep(1)

    pg.click(x=EXPRESS_COORDS["LOCAL_DEPRESSION"]["x"], y=EXPRESS_COORDS["LOCAL_DEPRESSION"]["y"]) # LOCAL_DEPRESSION.PNG
    pg.write("6", interval=0.01)

    pg.click(x=EXPRESS_COORDS["N-VALUE"]["x"], y=EXPRESS_COORDS["N-VALUE"]["y"]) # N_VALUE.PNG
    pg.write("0.013", interval=0.01)

    pg.click(x=EXPRESS_COORDS["COMPUTE_BY"]["x"], y=EXPRESS_COORDS["COMPUTE_BY"]["y"]) # COMPUTE_BY.PNG
    pg.click(x=EXPRESS_COORDS["KNOWN_Q"]["x"], y=EXPRESS_COORDS["KNOWN_Q"]["y"]) # KNOWN_Q_DROPDOWN.PNG
    time.sleep(1)


def input_values(slope, total_q, inlet):
    pg.click(x=EXPRESS_COORDS["NAME"]["x"], y=EXPRESS_COORDS["NAME"]["y"]) # EXPRESS_NAME.PNG
    pg.hotkey("ctrl", "a")
    pg.press("backspace")
    pg.write(inlet, interval=0.01)

    pg.click(x=EXPRESS_COORDS["SLOPE"]["x"], y=EXPRESS_COORDS["SLOPE"]["y"]) # SLOPE.PNG
    pg.write(str(slope), interval=0.01)

    pg.click(x=EXPRESS_COORDS["Q"]["x"], y=EXPRESS_COORDS["Q"]["y"]) # Q.PNG
    pg.write(str(total_q), interval=0.01)

    print(f"inputted values, slope: {slope}, q: {total_q}, inlet: {inlet}")


def run_express(inlet: str):
    pg.click(x=EXPRESS_COORDS["RUN"]["x"], y=EXPRESS_COORDS["RUN"]["y"]) # RUN.PNG

    save_hxp(inlet)
    time.sleep(0.5)

    save_pdf(inlet)
    time.sleep(0.5)

    save_csv(inlet)
    time.sleep(1)


def save_hxp(inlet: str):
    pg.click(x=EXPRESS_COORDS["FILE"]["x"], y=EXPRESS_COORDS["FILE"]["y"]) # FILE.PNG
    pg.click(x=EXPRESS_COORDS["SAVE_AS"]["x"], y=EXPRESS_COORDS["SAVE_AS"]["y"]) # SAVE_AS.PNG
    time.sleep(2)

    pg.write(inlet, interval=0.01)
    time.sleep(0.5)

    pg.press("enter")
    pg.press("enter")


def save_pdf(inlet: str):
    pg.click(x=EXPRESS_COORDS["PRINT"]["x"], y=EXPRESS_COORDS["PRINT"]["y"]) # PRINT.PNG
    pg.click(x=EXPRESS_COORDS["REPORT"]["x"], y=EXPRESS_COORDS["REPORT"]["y"]) # REPORT.PNG
    time.sleep(2)

    pg.write("Microsoft Print to PDF", interval=0.01)
    pg.press("enter")
    time.sleep(3)

    # pg.click(x=EXPRESS_COORDS["FILE_NAME"]["x"], y=EXPRESS_COORDS["FILE_NAME"]["y"])
    # time.sleep(0.5)

    pg.write(inlet, interval=0.01)
    time.sleep(0.5)

    pg.press("enter")
    time.sleep(3) # replace with PRINTING_COMPLETE.PNG

    pg.press("enter")


def save_csv(inlet: str):
    pg.click(x=EXPRESS_COORDS["FILE"]["x"], y=EXPRESS_COORDS["FILE"]["y"]) # FILE.PNG
    pg.click(x=EXPRESS_COORDS["EXPORT"]["x"], y=EXPRESS_COORDS["EXPORT"]["y"]) # EXPORT.PNG
    pg.click(x=EXPRESS_COORDS["RESULTS_GRID"]["x"], y=EXPRESS_COORDS["RESULTS_GRID"]["y"]) # RESULTS_GRID.PNG
    pg.click(x=EXPRESS_COORDS["CSV"]["x"], y=EXPRESS_COORDS["CSV"]["y"]) # CSV.PNG
    time.sleep(2)

    pg.write(inlet.strip() + ".csv", interval=0.01)
    time.sleep(0.5)

    pg.press("enter")


def read_csv(folder: str, inlet: str) -> pd.DataFrame:
    file = (inlet.strip()) + ".csv"
    path = Path(folder) / file

    df = pd.read_csv(path, header=1)
    df = df.loc[[1]].reset_index(drop=True)

    return df


def input_sag(row: Tuple[Cell, ...], prev_carryover_q: float):
    q = float(row[XL_INDEX["Q"]].value)

    total_q = q + prev_carryover_q
    if total_q >= 10.62:
        highlight_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        row[XL_INDEX["INTERCEPTED_FLOW"]].fill = highlight_fill

    row[XL_INDEX["INTERCEPTED_FLOW"]].value = total_q
    row[XL_INDEX["CARRYOVER_Q"]].value      = "N/A"
    row[XL_INDEX["SPREAD"]].value           = "N/A"
    row[XL_INDEX["DEPTH"]].value            = "N/A"


def edit_xl(row: Tuple[Cell, ...], df_express: pd.DataFrame) -> int:
    print(df_express)

    row[XL_INDEX["INTERCEPTED_FLOW"]].value = float(df_express.at[0, "Captured"])
    row[XL_INDEX["CARRYOVER_Q"]].value      = float(df_express.at[0, "Q"])
    row[XL_INDEX["SPREAD"]].value           = float(df_express.at[0, "Spread"])
    row[XL_INDEX["DEPTH"]].value            = float(df_express.at[0, "Depth"])

    if float(df_express.at[0, "Spread"]) > 6.2:
        highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        row[XL_INDEX["SPREAD"]].fill = highlight_fill

    carryover_q: float = float(df_express.at[0, "Q"])
    return carryover_q


def iter_xl(ws: Worksheet, folder):
    invalid = ["inlet #", "location", ""]
    prev_carryover_q = 0
    prev_to_inlet = ""

    for row in ws.iter_rows():

        inlet_cell = row[XL_INDEX["INLET"]]
        q_cell     = row[XL_INDEX["Q"]]
        long_cell  = row[XL_INDEX["LONG"]]
        type_cell  = row[XL_INDEX["ON-GRADE/SAG"]]

        type: str = type_cell.value
        inlet: str = inlet_cell.value

        if type == "Sag":
            input_sag(row, prev_carryover_q)
            continue

        skip_row = (inlet is None or inlet.lower() in invalid)
        if skip_row:
            prev_carryover_q = 0
            continue

        q = q_cell.value
        long_percent = long_cell.value * 100

        if prev_to_inlet == inlet:
            total_q = q + prev_carryover_q
        else:
            total_q = q


        input_values(long_percent, total_q, inlet)
        run_express(inlet)

        df_express       = read_csv(folder, inlet)
        prev_carryover_q = edit_xl(row, df_express)

        prev_to_inlet = inlet


def mkdirs(folder):
    dirs = ["/csvs", "/pdfs", "/hxps"]
    for dir in dirs:
        try:
            os.makedirs(folder+dir)
        except FileExistsError:
            os.rmdir(folder+dir)
            os.makedirs(folder+dir)


def move_files(folder):

    for file in os.listdir(folder):

        filename = os.fsdecode(file)
        original_path = Path(folder) / filename

        if filename.endswith(".pdf"):
            new_path = Path(folder) / "pdfs" / filename
            os.rename(original_path, new_path)

        elif filename.endswith(".hxp"):
            new_path = Path(folder) / "hxps" / filename
            os.rename(original_path, new_path)

        elif filename.endswith(".csv"):
            new_path = Path(folder) / "csvs" / filename
            os.rename(original_path, new_path)


if __name__ == "__main__":
    root = tk.Tk()
    root.withdraw()


    screen_w, screen_h = pg.size()
    if screen_w != 1920 or screen_h != 1080:
        raise RuntimeError("Screen resolution must be 1920x1080.")

    file_path = filedialog.askopenfilename(title="Select Gutter Spread Excel")
    folder = os.path.dirname(file_path)

    prepare_express(folder)

    wb = xl.load_workbook(file_path)
    ws = wb.worksheets[0]

    try:
        iter_xl(ws, folder)
        mkdirs(folder)
        move_files(folder)
    except Exception as e:
        pass

    path = Path(folder) / "UPDATED_GUTTER_SPREAD_EXCEL.xlsx"
    wb.save(path)